"""Excel file parser using openpyxl."""

from __future__ import annotations

import logging
from collections.abc import Iterator
from pathlib import Path
from typing import TYPE_CHECKING

from core.importer.parsers.base import ParsedFile
from core.importer.security import (
    ensure_import_row_limit,
    import_security_limits,
    normalize_header_cells,
    normalize_row_cells,
)

if TYPE_CHECKING:
    pass

logger = logging.getLogger(__name__)

# Supported file extensions
EXCEL_EXTENSIONS = {".xlsx"}


class ExcelParser:
    """Parser for Excel files (.xlsx).

    Attributes:
        sheet_name: Specific sheet to parse (None = first sheet).
        skip_rows: Number of rows to skip before header.
        max_rows: Maximum rows to read (None = all).
    """

    def __init__(
        self,
        sheet_name: str | None = None,
        skip_rows: int = 0,
        max_rows: int | None = None,
    ) -> None:
        """Initialize Excel parser.

        Args:
            sheet_name: Name of sheet to parse (None = first sheet).
            skip_rows: Number of rows to skip before reading.
            max_rows: Maximum number of data rows to read.
        """
        self.sheet_name = sheet_name
        self.skip_rows = skip_rows
        self.max_rows = max_rows

    def can_parse(self, path: Path) -> bool:
        """Check if this parser can handle the file."""
        return path.suffix.lower() in EXCEL_EXTENSIONS

    def parse(self, path: Path) -> ParsedFile:
        """Parse an Excel file and return headers and first 100 rows for preview."""
        if not path.exists():
            raise FileNotFoundError(f"File not found: {path}")

        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("openpyxl is required for Excel parsing.") from e

        limits = import_security_limits()
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheet = (
            self.sheet_name
            and workbook[self.sheet_name]
            or workbook.active
            or workbook.worksheets[0]
        )
        if len(workbook.sheetnames) > limits.max_sheets:
            workbook.close()
            raise ValueError(
                f"Workbook exceeds the maximum supported sheet count ({limits.max_sheets})."
            )
        sheet_name_used = sheet.title

        # Detect headers (find first non-empty row)
        headers = []
        header_row_idx = 1 + self.skip_rows

        for r_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx + 10, values_only=True)
        ):
            header_values = self._row_to_cells(row)
            if any(header_values):
                headers = normalize_header_cells(header_values)
                header_row_idx += r_idx
                break

        if not headers:
            max_columns = min(int(sheet.max_column or 0), limits.max_columns)
            headers = [f"Column_{i+1}" for i in range(max_columns)]

        # Estimate row count (openpyxl read_only might not have max_row)
        row_count = sheet.max_row - 1 - self.skip_rows if sheet.max_row else 0
        ensure_import_row_limit(max(0, int(row_count)))

        # Get preview rows
        preview_rows: list[dict[str, str]] = []
        it = self.iter_dicts(path)
        for row in it:
            preview_rows.append(row)
            if len(preview_rows) >= 100:
                break

        workbook.close()

        return ParsedFile(
            headers=headers,
            rows=preview_rows,
            source_type="excel",
            source_path=path,
            sheet_name=sheet_name_used,
            row_count=max(0, row_count),
        )

    def iter_dicts(self, path: Path) -> Iterator[dict[str, str]]:
        """Yield row dictionaries one by one."""
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = (
                self.sheet_name
                and workbook[self.sheet_name]
                or workbook.active
                or workbook.worksheets[0]
            )

            rows_iter = sheet.iter_rows(min_row=1 + self.skip_rows, values_only=True)

            # Find headers again (ensure consistency)
            headers_row = None
            for row in rows_iter:
                row_values = self._row_to_cells(row)
                if any(row_values):
                    headers_row = row_values
                    break

            if not headers_row:
                return

            headers = normalize_header_cells(headers_row)

            yielded = 0
            for row_values in rows_iter:
                normalized_values = normalize_row_cells(
                    self._row_to_cells(row_values),
                    expected_cols=len(headers),
                )
                # Skip completely empty rows
                if not any(normalized_values):
                    continue

                row_dict: dict[str, str] = {}
                for i, value in enumerate(normalized_values):
                    header = headers[i]
                    row_dict[header] = value
                yield row_dict
                yielded += 1
                if self.max_rows and yielded >= self.max_rows:
                    return
                ensure_import_row_limit(yielded)
        finally:
            workbook.close()

    def get_sheet_names(self, path: Path) -> list[str]:
        """Get list of sheet names in workbook.

        Args:
            path: Path to Excel file.

        Returns:
            List of sheet names.
        """
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True)
        names: list[str] = list(workbook.sheetnames)
        workbook.close()
        return names

    def _cell_to_string(self, value: object) -> str:
        """Convert cell value to string.

        Args:
            value: Cell value (can be None, int, float, str, datetime, etc.)

        Returns:
            String representation of the value.
        """
        if value is None:
            return ""
        if isinstance(value, bool):
            return "1" if value else "0"
        if isinstance(value, float):
            # Avoid scientific notation for large numbers
            if value == int(value):
                return str(int(value))
            return str(value)
        return str(value).strip()

    def _row_to_cells(self, row: tuple[object, ...] | list[object] | None) -> list[str]:
        if not row:
            return []
        return [self._cell_to_string(value) for value in row]
